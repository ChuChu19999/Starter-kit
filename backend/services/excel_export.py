import io
from datetime import date, datetime
from typing import Any, Awaitable, Callable, Optional
import pendulum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from core.logger import logger


def create_excel_file(
    headers: list[str],
    data: list[list[Any]],
    sheet_name: str = "Данные",
) -> io.BytesIO:
    """
    Создает Excel файл с заголовками и данными.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            data = []

        header_font = Font(bold=True, color="FFFFFF", name="Times New Roman")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Times New Roman")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                if value is None:
                    cell_value = ""
                else:
                    cell_value = value
                    if isinstance(value, datetime):
                        if value.tzinfo is not None:
                            dt_naive = value.astimezone().replace(tzinfo=None)
                        else:
                            dt_naive = value
                        dt_pendulum = pendulum.datetime(
                            dt_naive.year,
                            dt_naive.month,
                            dt_naive.day,
                            dt_naive.hour,
                            dt_naive.minute,
                            dt_naive.second,
                        )
                        cell_value = dt_pendulum.format("DD.MM.YYYY HH:mm:ss")
                    elif isinstance(value, date):
                        date_pendulum = pendulum.date(
                            value.year, value.month, value.day
                        )
                        cell_value = date_pendulum.format("DD.MM.YYYY")

                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.alignment = Alignment(horizontal="left", vertical="top")
                cell.font = data_font

        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(str(header))
            if ws.max_row > 1:
                for row in ws.iter_rows(
                    min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num
                ):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {str(e)}", exc_info=True)
        raise ValueError(f"Ошибка при сохранении Excel файла: {str(e)}")


def export_to_excel_from_dicts(
    headers: list[str],
    data: list[dict[str, Any]],
    sheet_name: str = "Данные",
    header_to_key_mapping: Optional[dict[str, str]] = None,
) -> io.BytesIO:
    """
    Создает Excel файл из списка словарей.
    """
    try:
        if header_to_key_mapping is None:
            header_to_key_mapping = {header: header for header in headers}

        excel_data = []
        for item in data:
            row = []
            for header in headers:
                key = header_to_key_mapping.get(header, header)
                value = item.get(key)
                row.append(value)
            excel_data.append(row)

        return create_excel_file(headers, excel_data, sheet_name)
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных в Excel: {str(e)}", exc_info=True)
        raise ValueError(f"Ошибка при экспорте данных в Excel: {str(e)}")


async def export_to_excel_async(
    headers: list[str],
    data_fetcher: Callable[..., Awaitable[list[Any]]],
    data_transformer: Callable[[Any], dict[str, Any]],
    sheet_name: str = "Данные",
    header_to_key_mapping: Optional[dict[str, str]] = None,
    **filter_kwargs: Any,
) -> io.BytesIO:
    """
    Универсальная функция для экспорта данных в Excel с поддержкой фильтров.
    """
    try:
        logger.info(
            f"Начало экспорта в Excel. Лист: {sheet_name}, " f"Фильтры: {filter_kwargs}"
        )

        raw_data = await data_fetcher(**filter_kwargs)

        data_dicts = []
        for idx, item in enumerate(raw_data):
            try:
                data_dicts.append(data_transformer(item))
            except Exception as e:
                logger.error(
                    f"Ошибка при преобразовании элемента {idx}: {str(e)}",
                    exc_info=True,
                )
                raise

        logger.info(f"Получено {len(data_dicts)} записей для экспорта")

        return export_to_excel_from_dicts(
            headers=headers,
            data=data_dicts,
            sheet_name=sheet_name,
            header_to_key_mapping=header_to_key_mapping,
        )
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных в Excel: {str(e)}", exc_info=True)
        raise ValueError(f"Ошибка при экспорте данных в Excel: {str(e)}")
